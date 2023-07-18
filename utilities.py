import pandas as pd
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

def excel_column_to_list(file_path, column_name):
    # 读取Excel文件
    df = pd.read_excel(file_path)

    # 获取指定列的数据
    column_data = df[column_name].tolist()

    # 在每个元素外添加单引号
    # column_data = ["'" + str(item) + "'" for item in column_data]

    return column_data


def get_urls(code, test_links=False):
    sina_url = f'https://finance.sina.com.cn/futures/quotes/{code}.shtml'
    shangjia_url = f'https://m.shangjia.com/qihuo/{code.lower()}/'

    if test_links:
        test_link(sina_url)
        test_link(shangjia_url)

    return sina_url, shangjia_url

def test_link(url):
    response = requests.head(url)

    if response.status_code != 200:
        print(f'Warning: {url} is not accessible. Status code: {response.status_code}')


def get_latest_price(code):
    sina_url, shangjia_url = get_urls(code)
    price = None

    # 创建一个chrome浏览器的驱动，设置为无头模式
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)

    # 尝试从新浪财经获取价格
    try:
        # 打开网页
        driver.get(sina_url)

        # 等待JavaScript加载完成
        driver.implicitly_wait(5)

        # 提取价格
        price_tag = driver.find_element(By.CSS_SELECTOR, 'td[class*="price"]')
        price = price_tag.text
        print(f'getting {code} price {price} succesfully from sina')
    except Exception as e:
        print(f'Error getting price from sina: {e}')

    # 如果从新浪财经获取价格失败，尝试从商家网获取价格
    if price is None or price == '--':
        try:
            # 打开网页
            driver.get(shangjia_url)

            # 等待JavaScript加载完成
            driver.implicitly_wait(10)

            # 提取价格
            price_tag = driver.find_element(By.CSS_SELECTOR, 'div[class*="remove_data"]')
            price = price_tag.text
            print(f'getting {code} price {price} succesfully from shangjia')
        except Exception as e:
            print(f'Error getting price from shangjia: {e}')

    # 不要忘记最后要关闭浏览器
    driver.quit()

    return price


def update_price(path, mode):
    # 加载工作簿
    wb = load_workbook(filename=path, data_only=True)

    # mode = 'work'
    # 待处理的sheet列表
    if mode == 'work':
        columns = ['A', 'C']
    else:
        columns = ['A', 'C', 'E']

    for column in columns:
        # 选择工作表
        ws = wb['Main']

        # 初始化行号
        row = 3

        # 循环处理每一行，直到A列没有数据
        while ws[f'{column}{row}'].value is not None:
            # 读取单元格
            code = ws[f'{column}{row}'].value
            # print(f'update {code} price')
            latest_price = get_latest_price(code)

            # 尝试将价格转换为浮点数
            try:
                price = float(latest_price)
            except ValueError:
                print(f'Warning: Cannot convert price "{latest_price}" to number. Please check the price.')
            else:
                # 计算右侧列名
                right_column = chr(ord(column) + 1)
                # 如果转换成功，更新价格
                ws[f'{right_column}{row}'].value = price

            # 处理下一行
            row += 1

            # 暂停一段时间
            time.sleep(5)  # 这里设置暂停一秒，可以根据实际需要调整

    # 保存修改
    wb.save(path)