from openpyxl import load_workbook
from get_latest_price import get_latest_price
import time

# 加载工作簿
wb = load_workbook(filename='Resource/target.xlsx', data_only=True)

mode = 'reset'
# 待处理的sheet列表
if mode == 'work':
    columns = ['A', 'C']
else:
    columns = ['A', 'C', 'E']

for column in columns:
    # 选择工作表
    ws = wb['Main']

    # 初始化行号
    row = 3

    # 循环处理每一行，直到A列没有数据
    while ws[f'{column}{row}'].value is not None:
        # 读取单元格
        code = ws[f'{column}{row}'].value
        print(f'update {code} price')
        latest_price = get_latest_price(code)

        # 尝试将价格转换为浮点数
        try:
            price = float(latest_price)
        except ValueError:
            print(f'Warning: Cannot convert price "{latest_price}" to number. Please check the price.')
        else:
            # 计算右侧列名
            right_column = chr(ord(column) + 1)
            # 如果转换成功，更新价格
            ws[f'{right_column}{row}'].value = price

        # 处理下一行
        row += 1

        # 暂停一段时间
        time.sleep(5)  # 这里设置暂停一秒，可以根据实际需要调整


# 保存修改
wb.save('Resource/target.xlsx')
